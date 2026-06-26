from __future__ import annotations

import argparse
import sys
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from monthly_order_ingestion.config import SHEET_TITLES, SheetKind
from monthly_order_ingestion.normalization import normalize_rows
from monthly_order_ingestion.staging_payload import write_staging_jsonl


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "office_rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for item in root.findall("main:si", NS):
        text_parts = [text.text or "" for text in item.findall(".//main:t", NS)]
        strings.append("".join(text_parts))
    return strings


def sheet_xml_path(archive: zipfile.ZipFile, sheet_title: str) -> str:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("rel:Relationship", NS)
    }
    for sheet in workbook.findall("main:sheets/main:sheet", NS):
        if sheet.attrib.get("name") != sheet_title:
            continue
        rel_id = sheet.attrib[f"{{{NS['office_rel']}}}id"]
        target = rel_targets[rel_id]
        return target if target.startswith("xl/") else f"xl/{target}"
    raise ValueError(f"sheet not found in workbook.xml: {sheet_title}")


def raw_cell_values(path: Path, sheet_title: str) -> dict[str, str]:
    with zipfile.ZipFile(path) as archive:
        strings = shared_strings(archive)
        sheet_path = sheet_xml_path(archive, sheet_title)
        root = ET.fromstring(archive.read(sheet_path))
    values: dict[str, str] = {}
    for cell in root.findall(".//main:c", NS):
        ref = cell.attrib.get("r")
        if not ref:
            continue
        cell_type = cell.attrib.get("t")
        value_node = cell.find("main:v", NS)
        if cell_type == "inlineStr":
            text_parts = [text.text or "" for text in cell.findall(".//main:t", NS)]
            values[ref] = "".join(text_parts)
        elif value_node is not None:
            value = value_node.text or ""
            if cell_type == "s":
                values[ref] = strings[int(value)]
            else:
                values[ref] = value
    return values


def worksheet_values(path: Path, sheet_title: str) -> list[list[object]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_title not in workbook.sheetnames:
        raise ValueError(f"sheet not found: {sheet_title}")
    worksheet = workbook[sheet_title]
    raw_values = raw_cell_values(path, sheet_title)
    rows: list[list[object]] = []
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        for row_number, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            row = list(raw_row)
            for column_index, value in enumerate(row, start=1):
                if value == "#VALUE!":
                    ref = f"{get_column_letter(column_index)}{row_number}"
                    row[column_index - 1] = raw_values.get(ref, value)
            while row and row[-1] is None:
                row.pop()
            rows.append(row)
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert one xlsx sheet tab to staging JSONL.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--source", required=True)
    parser.add_argument("--sheet-kind", required=True, choices=[kind.value for kind in SheetKind])
    parser.add_argument("--source-file-id", required=True)
    parser.add_argument("--source-file-name", required=True)
    parser.add_argument("--source-yyyymm", required=True)
    parser.add_argument("--drive-modified-time", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    kind = SheetKind(args.sheet_kind)
    values = worksheet_values(Path(args.xlsx), SHEET_TITLES[kind])
    rows = normalize_rows(
        values,
        source=args.source,
        sheet_kind=kind.value,
        source_file_id=args.source_file_id,
        source_file_name=args.source_file_name,
        source_yyyymm=args.source_yyyymm,
        drive_modified_time=args.drive_modified_time,
    )
    row_count = write_staging_jsonl(rows, args.output)
    print(f"wrote {row_count} rows to {args.output}")


if __name__ == "__main__":
    main()
